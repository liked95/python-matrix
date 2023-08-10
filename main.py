import openpyxl

# 1. Swap function to swap row i and j from the given matrix
def swap(matrix, i, j):
    matrix[i], matrix[j] = matrix[j], matrix[i]
    return matrix

# 2. Scale function to multiple a row by certain number
def scale(matrix, i, j):
    # Loop through each element (or column) in the ith row and multiple it by j
    for column in range(len(matrix[i])):
        matrix[i][column] *= j
    return matrix

# 3. Combine function to multiple the content of ith row by k and add the result to row j
def row_combine(matrix, i, j, k):
    # Create an empty array to store value of row i * k
    array = []
    for col in range(len(matrix[i])):
        array.append(matrix[i][col] * k)

    # Loop row j and increment each col by the corresponding col from the newly created array
    for col in range(len(array)):
        matrix[j][col] += array[col]

    return matrix

# 4. Function to read matrix from an Excel file
def read_matrix_from_excel(file_path, sheet_name, start_cell, end_cell):
    # Load the excel file and assign to workbook and then assign the sheet in variable sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    matrix = []

    # Create an empty array matrix and then loop through the sheet rows to write the value to the matrix
    for row in sheet.iter_rows(min_row=start_cell[0], max_row=end_cell[0], min_col=start_cell[1], max_col=end_cell[1]):
        matrix.append([cell.value for cell in row])

    return matrix


# 6. Function to read matrix from an Excel file
def write_matrix_to_excel(matrix, file_path, sheet_name, start_cell):
    # Load the excel file and assign to workbook and then assign the sheet in variable sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # This is the reverse version from read_matrix_from_excel, which write matrix values to excel sheet and then save the work book
    for i, row in enumerate(matrix):
        for j, value in enumerate(row):
            sheet.cell(row=start_cell[0] + i,
                       column=start_cell[1] + j, value=value)

    workbook.save(file_path)


# Main function
def echelon(file_path):
    matrix = read_matrix_from_excel(file_path, 'Sheet1', (1, 1), (5, 5))

    # Loop through each row of the matrix in diagonal direction
    for i in range(len(matrix)):
        # If the entry is not equal to 0, scale that row with the inverse value of that value (1/matrix[i][i])
        if matrix[i][i] != 0:
            # This is step a
            matrix = scale(matrix, i, 1 / matrix[i][i])
            # Run through a loop with j from i+1 to the final row using row_combine to subtract matrix(j,i) times row i from row j
            for j in range(i + 1, len(matrix)):
                matrix = row_combine(matrix, i, j, -matrix[j][i])
            # If the entry is zero, go to this branch
        else:
            # Create a var to track if there exists a non zero row
            non_zero_row = None
            # loop through the rows below row i, at col i and check if the cell is 0.
            # Find a non-zero row to swap with if diagonal entry is zero
            for j in range(i + 1, len(matrix)):
                if matrix[j][i] != 0:
                    non_zero_row = j
                    break

            if non_zero_row is not None:
                # Swap the current row with the non-zero row
                matrix = swap(matrix, i, non_zero_row)
                # Repeat step a
                matrix = scale(matrix, i, 1 / matrix[i][i])
                for j in range(i + 1, len(matrix)):
                    matrix = row_combine(matrix, i, j, -matrix[j][i])

    # Write the final value of matrix to cells F1:J5 in the original Excel
    write_matrix_to_excel(matrix, file_path, 'Sheet1', (1, 6))


# Call the echelon function with the input Excel file path
echelon('echelon-input.xlsx')
